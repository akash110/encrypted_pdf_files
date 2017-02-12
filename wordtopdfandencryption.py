#module to convert the word files to pdf  files 
import PyPDF2
import os
import  openpyxl
from win32com import client

folder = "C:\Users\Akash.Rai\Desktop\work_final\word_files"
file_type = 'docx'
out_folder = "C:\Users\Akash.Rai\Desktop\work_final\converted_pdf_files"

if not os.path.exists(out_folder):
    os.makedirs(out_folder)

os.chdir(folder)

try:
    word = client.DispatchEx("Word.Application")
    for files in os.listdir("."):
        if files.endswith(".docx") or files.endswith('doc'):
            out_name = files.replace(file_type, r"pdf")
            in_file = os.path.abspath(folder + "\\" + files)
            out_file = os.path.abspath(out_folder + "\\" + out_name)
            doc = word.Documents.Open(in_file)
            doc.SaveAs(out_file, FileFormat=17)
            doc.Close()
except Exception, e:
    print e
finally:
    word.Quit()
    
    
#extracting password from excel files 
wb = openpyxl.load_workbook ('C:\Users\Akash.Rai\Desktop\work_final\excel file\password.xlsx', data_only = True)
sheet = wb.get_sheet_by_name('Sheet1')

print sheet.max_row
Password_List = []
for i in range(2, sheet.max_row+1):
  Password_List.append(sheet.cell(row=i, column=5).value.encode('utf-8'))  
    
print Password_List


#module to convert converted pdf files to encrypted pfd files 
    
files_list=[]
path='C:\Users\Akash.Rai\Desktop\work_final\converted_pdf_files'
for files in os.listdir(path):
   if files.endswith(('.PDF', '.pdf', '.Pdf')):
    files_list.append(files)
    
print files_list
    
for file_name,password in zip(files_list,Password_List):
        str1=path + "\\" +file_name
        pdf1File = open(str1, 'rb')
        pdf1Reader = PyPDF2.PdfFileReader(pdf1File)
        pdfWriter = PyPDF2.PdfFileWriter()

        for pageNum in range(pdf1Reader.numPages):
            pageObj = pdf1Reader.getPage(pageNum)
            pdfWriter.addPage(pageObj)

        pdfWriter.encrypt(password)
        ## Please create a seperate folder
        encrypted_folder='C:\Users\Akash.Rai\Desktop\work_final\encrypted_pdf_files'
        if not os.path.exists(encrypted_folder):
            os.makedirs(encrypted_folder)
        output_file_name = os.path.join('C:\Users\Akash.Rai\Desktop\work_final\encrypted_pdf_files', file_name)
        pdfOutputFile = open(output_file_name, 'wb')
        pdfWriter.write(pdfOutputFile)
        pdfOutputFile.close()
        pdf1File.close()
        
